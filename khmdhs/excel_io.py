"""Reading the source xlsx and writing the enriched copy."""
from __future__ import annotations

import logging
import sqlite3
from copy import copy
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from khmdhs.config import NEW_HEADERS


def read_adams(xlsx_path: Path) -> list[str]:
    """Return the ADAM codes from column B of the input file (rows 2..N)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    adams: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 2:
            continue
        val = row[1]
        if val is None:
            continue
        s = str(val).strip()
        if len(s) >= 10 and s[:2].isdigit() and any(c.isalpha() for c in s[2:6]):
            adams.append(s)
        else:
            logging.warning("Row %d column B is not an ADAM: %r — skipping.", row_idx, s)
    wb.close()
    return adams


def _join(seq: Iterable[Any], sep: str = " | ") -> str:
    return sep.join(str(x) for x in seq if x is not None and str(x) != "")


def _enrichment_for(conn: sqlite3.Connection, adam: str) -> list[Any]:
    """Build the 8 cell values appended after column H for one ADAM."""
    row = conn.execute(
        """SELECT procedure_type, units_operator_name, signer_name,
                  nuts_code, nuts_region_name, nuts_city, nuts_postal_code,
                  total_cost_with_vat
             FROM contracts WHERE reference_number = ?""",
        (adam,),
    ).fetchone()
    if row is None:
        return ["", "", "", "", "", "", "", ""]

    procedure, units_name, signer_name, nuts_code, nuts_region, nuts_city, nuts_zip, cost_with_vat = row

    contractor_names = [r[0] for r in conn.execute(
        "SELECT name FROM contractors WHERE reference_number = ? ORDER BY seq", (adam,))]
    contractor_vats = [r[0] for r in conn.execute(
        "SELECT vat_number FROM contractors WHERE reference_number = ? ORDER BY seq", (adam,))]
    cpv_codes = [r[0] for r in conn.execute(
        "SELECT cpv_code FROM contract_cpvs WHERE reference_number = ? ORDER BY seq", (adam,))]

    nuts_parts = []
    if nuts_code or nuts_region:
        nuts_parts.append(_join([nuts_code, nuts_region], sep=" — "))
    if nuts_city or nuts_zip:
        nuts_parts.append(_join([nuts_city, nuts_zip], sep=", "))
    nuts_text = " · ".join(p for p in nuts_parts if p)

    return [
        _join(contractor_names),
        _join(contractor_vats),
        procedure or "",
        units_name or "",
        nuts_text,
        signer_name or "",
        cost_with_vat if cost_with_vat is not None else "",
        _join(cpv_codes),
    ]


def write_enriched_xlsx(conn: sqlite3.Connection, src_xlsx: Path, dst_xlsx: Path) -> None:
    """Open the source workbook in memory, append NEW_HEADERS columns, save to dst."""
    dst_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(src_xlsx)
    ws = wb.active
    start_col = ws.max_column + 1
    header_template = ws.cell(row=1, column=1)
    for offset, header in enumerate(NEW_HEADERS):
        cell = ws.cell(row=1, column=start_col + offset, value=header)
        cell.font = copy(header_template.font)
        cell.fill = copy(header_template.fill)
        cell.alignment = copy(header_template.alignment)

    for row_idx in range(2, ws.max_row + 1):
        adam_cell = ws.cell(row=row_idx, column=2).value
        if adam_cell is None:
            continue
        adam = str(adam_cell).strip()
        for offset, val in enumerate(_enrichment_for(conn, adam)):
            ws.cell(row=row_idx, column=start_col + offset, value=val)

    for offset, header in enumerate(NEW_HEADERS):
        col_letter = ws.cell(row=1, column=start_col + offset).column_letter
        ws.column_dimensions[col_letter].width = max(20, min(60, len(header) + 4))

    wb.save(dst_xlsx)
    wb.close()
